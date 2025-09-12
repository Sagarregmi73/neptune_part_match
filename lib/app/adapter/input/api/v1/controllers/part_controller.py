# lib/app/adapter/input/api/v1/controllers/part_controller.py
from fastapi import APIRouter, HTTPException,UploadFile,File
from typing import List
from lib.app.application.use_cases.crud_part_usecase import CrudPartUseCase
from lib.app.domain.dtos.part_number_dto import PartNumberDTO
from lib.app.domain.entities.part_number import PartNumber
from openpyxl import load_workbook
from io import BytesIO,StringIO
from lib.core.aws.s3_client import s3, S3_BUCKET_NAME  # your S3 client
from lib.core.aws.neptune_bulk_loader import trigger_bulk_load
from lib.core.utils.config_loader import NEPTUNE_IAM_ROLE_ARN  # env variable for Neptune
import csv
import time
import requests
router = APIRouter()

# Inject your repository here (NeptuneRepository/PostgresRepository/MongoRepository)
from lib.core.utils.container import get_part_repository
usecase = CrudPartUseCase(get_part_repository())

@router.post("/", response_model=PartNumberDTO)
def create_part(part_dto: PartNumberDTO):
    part = PartNumber(part_dto.part_number, part_dto.specs, part_dto.notes)
    created = usecase.create_part(part)
    return PartNumberDTO(**vars(created))

@router.get("/{part_number}", response_model=PartNumberDTO)
def get_part(part_number: str):
    part = usecase.get_part(part_number)
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    return PartNumberDTO(**vars(part))

@router.put("/{part_number}", response_model=PartNumberDTO)
def update_part(part_number: str, part_dto: PartNumberDTO):
    part = PartNumber(part_number, part_dto.specs, part_dto.notes)
    updated = usecase.update_part(part)
    return PartNumberDTO(**vars(updated))

@router.delete("/{part_number}")
def delete_part(part_number: str):
    success = usecase.delete_part(part_number)
    return {"success": success}

@router.get("/", response_model=List[PartNumberDTO])
def list_parts():
    parts = usecase.list_parts()
    return [PartNumberDTO(**vars(p)) for p in parts]

@router.post("/upload")
async def upload_parts(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only XLSX files are supported")
    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        # use newline="" to avoid blank lines on some platforms
        vertices_output = StringIO(newline="")
        edges_output = StringIO(newline="")

        vertices_writer = csv.writer(vertices_output)
        edges_writer = csv.writer(edges_output)

        # Headers
        vertices_writer.writerow(["~id", "~label", "part_number", "specs", "notes"])
        edges_writer.writerow(["~id", "~from", "~to", "~label", "match_type"])

        seen_vertices = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            (in_part, in_specs, in_notes, out_part, out_specs, out_notes, match_type) = row

            # Vertices
            if in_part and in_part not in seen_vertices:
                vertices_writer.writerow([in_part, "PartNumber", in_part, in_specs or "", in_notes or ""])
                seen_vertices.add(in_part)
            if out_part and out_part != "-" and out_part not in seen_vertices:
                vertices_writer.writerow([out_part, "PartNumber", out_part, out_specs or "", out_notes or ""])
                seen_vertices.add(out_part)

            # Edges
            if out_part and out_part != "-" and match_type in ("Perfect", "Partial"):
                edge_id = f"{in_part}_to_{out_part}"
                edges_writer.writerow([edge_id, in_part, out_part, "replacement", match_type])

        vertices_output.seek(0)
        edges_output.seek(0)

        base_name = file.filename.replace(".xlsx", "")
        vertices_key = f"uploads/{base_name}_vertices.csv"
        edges_key = f"uploads/{base_name}_edges.csv"

        # Upload and confirm via head_object
        s3.put_object(Bucket=S3_BUCKET_NAME, Key=vertices_key, Body=vertices_output.getvalue())
        s3.put_object(Bucket=S3_BUCKET_NAME, Key=edges_key, Body=edges_output.getvalue())

        # Confirm objects exist (retry a few times)
        def wait_for_s3(bucket, key, retries=5, backoff=1.0):
            for i in range(retries):
                try:
                    s3.head_object(Bucket=bucket, Key=key)
                    return True
                except Exception:
                    time.sleep(backoff * (i + 1))
            return False

        if not wait_for_s3(S3_BUCKET_NAME, vertices_key):
            raise HTTPException(status_code=500, detail=f"Uploaded vertices file not visible: {vertices_key}")
        if not wait_for_s3(S3_BUCKET_NAME, edges_key):
            raise HTTPException(status_code=500, detail=f"Uploaded edges file not visible: {edges_key}")

        # Debug prints (view in docker logs)
        print("Uploaded files to S3:")
        print(f"s3://{S3_BUCKET_NAME}/{vertices_key}")
        print(f"s3://{S3_BUCKET_NAME}/{edges_key}")
        print("Triggering Neptune loader for vertices then edges...")

        try:
            loader_response_vertices = trigger_bulk_load(
                s3_path=f"s3://{S3_BUCKET_NAME}/{vertices_key}",
                iam_role_arn=NEPTUNE_IAM_ROLE_ARN
            )
        except Exception as e:
            # attach loader error details
            print("Vertices loader failed:", str(e))
            raise HTTPException(status_code=500, detail=f"Vertices loader failed: {str(e)}")

        try:
            loader_response_edges = trigger_bulk_load(
                s3_path=f"s3://{S3_BUCKET_NAME}/{edges_key}",
                iam_role_arn=NEPTUNE_IAM_ROLE_ARN
            )
        except Exception as e:
            print("Edges loader failed:", str(e))
            raise HTTPException(status_code=500, detail=f"Edges loader failed: {str(e)}")

        return {
            "message": "Upload successful",
            "vertices_file": vertices_key,
            "edges_file": edges_key,
            "loader_response_vertices": loader_response_vertices,
            "loader_response_edges": loader_response_edges
        }

    except HTTPException:
        raise
    except Exception as exc:
        print("Unhandled error in upload_parts:", repr(exc))
        raise HTTPException(status_code=500, detail=f"Internal error: {str(exc)}")